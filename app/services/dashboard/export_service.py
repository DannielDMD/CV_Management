from typing import Optional
from io import BytesIO
import pandas as pd
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import extract
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

from app.models.candidato_model import Candidato
from app.models.catalogs.ciudad import Ciudad
from app.models.educacion_model import Educacion
from app.models.experiencia_model import ExperienciaLaboral
from app.models.conocimientos_model import CandidatoConocimiento
from app.models.preferencias import PreferenciaDisponibilidad

def exportar_candidatos_detallados_excel(db: Session, año: Optional[int] = None) -> BytesIO:
    """
    Genera un archivo Excel con toda la información detallada de cada candidato.
    Si se especifica un año, solo se exportan los registrados ese año.
    """

    # 1. Consultar candidatos con joins
    query = (
        db.query(Candidato)
        .options(
            joinedload(Candidato.ciudad).joinedload(Ciudad.departamento),
            joinedload(Candidato.ciudad),
            joinedload(Candidato.cargo),
            joinedload(Candidato.centro_costos),
            joinedload(Candidato.motivo_salida),
            joinedload(Candidato.educaciones).joinedload(Educacion.nivel_educacion),
            joinedload(Candidato.educaciones).joinedload(Educacion.titulo),
            joinedload(Candidato.educaciones).joinedload(Educacion.institucion),
            joinedload(Candidato.educaciones).joinedload(Educacion.nivel_ingles),
            joinedload(Candidato.experiencias).joinedload(ExperienciaLaboral.rango_experiencia),
            joinedload(Candidato.conocimientos).joinedload(CandidatoConocimiento.habilidad_blanda),
            joinedload(Candidato.conocimientos).joinedload(CandidatoConocimiento.habilidad_tecnica),
            joinedload(Candidato.conocimientos).joinedload(CandidatoConocimiento.herramienta),
            joinedload(Candidato.preferencias).joinedload(PreferenciaDisponibilidad.disponibilidad),
            joinedload(Candidato.preferencias).joinedload(PreferenciaDisponibilidad.rango_salarial),
            joinedload(Candidato.preferencias).joinedload(PreferenciaDisponibilidad.motivo_salida),
        )
    )
    if año:
        query = query.filter(extract("year", Candidato.fecha_registro) == año)

    candidatos = query.all()

    # 2. Construir registros
    registros = []
    candidatos.sort(key=lambda c: c.fecha_registro)
    for idx, c in enumerate(candidatos, start=1):

        educ = c.educaciones[0] if c.educaciones else None
        exp = c.experiencias[0] if c.experiencias else None
        pref = c.preferencias[0] if c.preferencias else None

        hb = [ci.habilidad_blanda.nombre_habilidad_blanda for ci in c.conocimientos if ci.tipo_conocimiento == "blanda" and ci.habilidad_blanda]
        ht = [ci.habilidad_tecnica.nombre_habilidad_tecnica for ci in c.conocimientos if ci.tipo_conocimiento == "tecnica" and ci.habilidad_tecnica]
        hr = [ci.herramienta.nombre_herramienta for ci in c.conocimientos if ci.tipo_conocimiento == "herramienta" and ci.herramienta]

        registros.append({
            "#": idx,
            "ID del Candidato": c.id_candidato,
            "Nombre Completo": c.nombre_completo,
            "Correo Electrónico": c.correo_electronico,
            "CC": c.cc,
            "Fecha de Nacimiento": c.fecha_nacimiento,
            "Teléfono": c.telefono,
            "Departamento de Residencia": c.ciudad.departamento.nombre_departamento if c.ciudad and c.ciudad.departamento else None,
            "Ciudad/Municipio": c.ciudad.nombre_ciudad if c.ciudad else None,
            "Descripción del Perfil": c.descripcion_perfil,
            "Cargo de Interés": c.cargo.nombre_cargo if c.cargo else None,
            "Nombre (Otro Cargo)": c.nombre_cargo_otro,
            "¿Traba Actualemente en Joyco?": c.trabaja_actualmente_joyco,
            "Centro de Costos": c.centro_costos.nombre_centro_costos if c.centro_costos else None,
            "Nombre (Otro Centro de Costos)": c.nombre_centro_costos_otro,
            "¿Ha Trabajado en Joyco?": c.ha_trabajado_joyco,
            "Motivo de Salida": c.motivo_salida.descripcion_motivo if c.motivo_salida else None,
            "Nombre (Otro Motivo de Salida)": c.otro_motivo_salida,
            "Tiene Referido": c.tiene_referido,
            "Nombre del Referido": c.nombre_referido,
            "Estado": c.estado,
            "¿Formulario Completo?": c.formulario_completo,
            "¿Acpetó Política de Datos?": c.acepta_politica_datos,
            # Educación
            "Ultimo Nivel Educativo": educ.nivel_educacion.descripcion_nivel if educ else None,
            "Título Obtenido": educ.titulo.nombre_titulo if educ and educ.titulo else None,
            "Nombre (Otro Título)": educ.nombre_titulo_otro if educ else None,
            "Institución Académica": educ.institucion.nombre_institucion if educ and educ.institucion else None,
            "Nombre (Otro Institución Académica)": educ.nombre_institucion_otro if educ else None,
            "Año de Graduación": educ.anio_graduacion if educ else None,
            "Nivel de Inglés": educ.nivel_ingles.nivel if educ and educ.nivel_ingles else None,
            # Experiencia
            "Experiencia Laboral": exp.rango_experiencia.descripcion_rango if exp else None,
            "Última Empresa": exp.ultima_empresa if exp else None,
            "Último Cargo": exp.ultimo_cargo if exp else None,
            "Funciones Relizadas": exp.funciones if exp else None,
            "Desde": exp.fecha_inicio if exp else None,
            "Hasta": exp.fecha_fin if exp else None,
            # Conocimientos
            "Habilidades Blandas": ", ".join(hb),
            "Habilidades Técnicas": ", ".join(ht),
            "Herramientas": ", ".join(hr),
            # Preferencias
            "¿Disponibilidad de Viajar?": pref.disponibilidad_viajar if pref else None,
            "¿Disponibilidad de Inicio?": pref.disponibilidad.descripcion_disponibilidad if pref else None,
            "Pretensión Salarial": pref.rango_salarial.descripcion_rango if pref else None,
            "¿Trabaja Actualmente?": pref.trabaja_actualmente if pref else None,
            "Motivo de Salida": pref.motivo_salida.descripcion_motivo if pref and pref.motivo_salida else None,
            "Nombre (Otro Motivo de Salida (Preferencias))": pref.otro_motivo_salida if pref else None,
            "Razón para Trabajar en Joyco": pref.razon_trabajar_joyco if pref else None,
            # Fecha al final
            "Fecha de Registro": c.fecha_registro,
        })

    df = pd.DataFrame(registros)

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidatos"

    if not df.empty:
        # Ordenar y mover columna
        df.sort_values(by="Fecha de Registro", ascending=True, inplace=True)
        columnas_ordenadas = [col for col in df.columns if col != "Fecha de Registro"] + ["Fecha de Registro"]
        df = df[columnas_ordenadas]

        # Escribir filas al worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            ws.append(row)
            for cell in ws[r_idx]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Ajuste de anchos
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        # Estilo de tabla
        tab = Table(displayName="TablaCandidatos", ref=ws.dimensions)
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    else:
        # Sin datos → mostrar mensaje
        ws.append(["Sin datos disponibles para el año seleccionado."])
        ws.column_dimensions["A"].width = 50
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

